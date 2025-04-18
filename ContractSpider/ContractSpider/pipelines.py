
from datetime import datetime
from scrapy.exceptions import DropItem


import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional

from ContractSpider.utils.excel_writer import append_df_to_excel, has_data_in_sheet


class ContractPipeline:

    def __init__(self):
        # 存储目录
        self.base_folder = "downloads"
        os.makedirs(self.base_folder, exist_ok=True)

    def process_item(self, item, spider):
        spider.custom_logger.info(f"收到合同数据: {item}")
        file_path = item.get("file_path")
        if not file_path:
            spider.custom_logger.error("缺少文件路径，跳过保存")
            return item  # 跳过无效数据

        # 将数据转换为 DataFrame
        data = {
            "签订日期": [item["sign_date"]],
            "发布时间": [item["publish_date"]],
            "采购人": [item["purchaser"]],
            "供应商": [item["supplier"]],
            "代理机构": [item["agent"]],
            "合同名称": [item["contract_name"]],
            "项目名称": [item["project_name"]],
            "网页链接": [item["contract_link"]],
        }

        df = pd.DataFrame(data)

        # 获取 pandas 版本
        pandas_version = pd.__version__

        header_needed = not has_data_in_sheet(file_path, sheet_name="Contracts")

        # 根据 pandas 版本选择写入方式
        if self.is_pandas_version_less_than("1.4.0", pandas_version):
            append_df_to_excel(file_path, df, sheet_name="Contracts", header=header_needed)
        else:
            self.append_data_to_excel(file_path, df)

        spider.custom_logger.info(f"保存合同数据: {file_path}")
        return item

    def is_pandas_version_less_than(self, version_str: str, current_version: str) -> bool:
        """
        Helper function to compare pandas version.
        """
        return tuple(map(int, current_version.split('.'))) < tuple(map(int, version_str.split('.')))

    def append_data_to_excel(self, filename: Path, df: pd.DataFrame, sheet_name: str = 'Contracts',
                             startrow: Optional[int] = None):
        """
        Append data to an existing Excel file using pd.ExcelWriter.
        """
        file_exists = os.path.exists(filename)

        if file_exists:
            # 打开现有的 Excel 文件
            book = load_workbook(filename)
            sheet = book["Contracts"] if "Contracts" in book.sheetnames else None

            if sheet:
                # 获取当前sheet中的最大行数
                startrow = sheet.max_row
                print(f"开始写入行号: {startrow}")
            else:
                startrow = 0  # 如果没有 Contracts sheet，重新开始
                print('没有 Contracts sheet，重新开始')

            with pd.ExcelWriter(filename, mode="a", engine="openpyxl") as writer:
                # 将数据写入指定位置
                df.to_excel(writer, index=False, sheet_name="Contracts", header=False, startrow=startrow)
        else:
            # 如果文件不存在，创建一个新文件并写入数据
            df.to_excel(filename, index=False, sheet_name="Contracts")



import pandas as pd
from packaging import version
from openpyxl import load_workbook
from datetime import datetime
import os
from scrapy.exceptions import DropItem

class DetailPipeline:
    def __init__(self):
        self.base_folder = "detail_downloads"
        os.makedirs(self.base_folder, exist_ok=True)

        self.headers_map = {
            "contract_number": "合同编号",
            "contract_name": "合同名称",
            "project_number": "项目编号",
            "project_name": "项目名称",
            "purchaser": "采购人（甲方）",
            "purchaser_address": "采购人地址",
            "purchaser_contact": "采购人联系方式",
            "supplier": "供应商（乙方）",
            "supplier_address": "供应商地址",
            "supplier_contact": "供应商联系方式",
            "main_product_name": "主要标的名称",
            "specifications": "规格型号（或服务要求）",
            "quantity": "主要标的数量",
            "unit_price": "主要标的单价",
            "contract_amount": "合同金额",
            "performance_location": "履约地点",
            "procurement_method": "采购方式",
            "contract_sign_date": "合同签订日期",
            "contract_announcement_date": "合同公告日期",
            "attachment_name": "附件名称",
            "attachment_download_url": "附件下载链接"
        }

    def process_item(self, item, spider):
        # spider.custom_logger.info(f"[DetailPipeline] 接收详情数据: {item}")

        # 校验和解析公告日期
        announce_date_str = item.get("contract_announcement_date", "")
        if not announce_date_str:
            raise DropItem("缺少合同公告日期")
        try:
            announce_date = datetime.strptime(announce_date_str, "%Y-%m-%d")
        except ValueError:
            raise DropItem(f"合同公告日期格式错误: {announce_date_str}")

        # 构建保存路径
        folder = announce_date.strftime("%Y-%m")
        folder_path = os.path.join(self.base_folder, folder)
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, f"{announce_date.strftime('%Y-%m-%d')}.xlsx")

        # 附件字段转为字符串
        if isinstance(item.get("attachment_name"), list):
            item["attachment_name"] = ", ".join(item["attachment_name"])
        if isinstance(item.get("attachment_download_url"), list):
            item["attachment_download_url"] = ", ".join(item["attachment_download_url"])

        # 映射字段生成 DataFrame
        item_dict = {
            self.headers_map[k]: v for k, v in dict(item).items() if k in self.headers_map
        }
        df = pd.DataFrame([item_dict])

        # 判断是否存在表头
        sheet_name = "Details"
        file_exists = os.path.exists(file_path)
        sheet_has_data = has_data_in_sheet(file_path, sheet_name=sheet_name)

        header = not sheet_has_data  # 有数据则不写表头

        # Pandas 版本判断
        if version.parse(pd.__version__) < version.parse("1.4.0"):
            append_df_to_excel(file_path, df, sheet_name=sheet_name, header=header)
        else:
            # 高版本直接使用 overlay
            with pd.ExcelWriter(file_path, mode="a" if file_exists else "w",
                                engine="openpyxl", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, header=header, sheet_name=sheet_name)

        return item


